# %%
import concurrent.futures
import openpyxl
import requests
import re
import html
from bs4 import BeautifulSoup
from datetime import datetime

# %%

def fetch_bf_spec_data(bf_spec_no):
    url = "https://www.g2b.go.kr/pn/pnz/pnza/BfSpec/selectBfSpec.do"
    
    payload = {
        "dlParamM": {
            "bfSpecRegNo": bf_spec_no,
            "prcmBsneSeCd": "",  # 필요에 따라 수정
            "opnnSqno": "",      # 필요에 따라 수정
            "jobType": ""        # 필요에 따라 수정
        }
    }

    headers = {
        'Content-Type': 'application/json;charset=UTF-8',
        'Accept': 'application/json',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        'Referer': 'https://www.g2b.go.kr/',
        'Origin': 'https://www.g2b.go.kr',
        'Cookie': 'WHATAP=z1c19kr5h17kv5; XTVID=A2501060840080519; JSESSIONID=NTZjMjBjMTAtZTg5YS00OGVlLTkzNTktZGNiYjIwMjY0ZmZl;',  # 최신 쿠키 입력
        'Menu-Info': '{"menuNo":"01141","menuCangVal":"PRVA004_02","bsneClsfCd":"%EC%97%85130025","scrnNo":"05929"}',
        'SubmissionId': 'mf_wfm_container_sbmSrch',
        'Priority': 'u=1, i',
        'Sec-CH-UA': '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
        'Sec-CH-UA-Mobile': '?0',
        'Sec-CH-UA-Platform': '"Windows"',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
    }

    response = requests.post(url, headers=headers, json=payload)

    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error: {response.status_code} {response.text}")
        return None

"""
def process_row(row, worksheet):
    bf_spec_no = worksheet.cell(row=row, column=11).value
    data = fetch_bf_spec_data(bf_spec_no)

    if data:
        # 데이터 기록
        # 필요한 정보 추출
        dlBfSpecM = data.get("dlBfSpecM", {})
        budget_amount = dlBfSpecM.get("alotBgtPrspAmt", "N/A")
        worksheet.cell(row=row, column=7).value = budget_amount  # 배정예산액 (부가세 포함)
        print(f"Row {row}: Data written successfully -> {dlBfSpecM}")
    else:
        print(f"Row {row}: No data extracted for bfSpecNo {bf_spec_no}")

    return budget_amount
"""
# %%
"""
def process_row(row, worksheet):
    organization_name = worksheet.cell(row=row, column=14).value
    
    if organization_name is None:
        return

    
    budget_amount = process_row(row, worksheet)
    
    if budget_amount is None:
        worksheet.cell(row=row, column=6).value = "?"
        worksheet.cell(row=row, column=7).value = "예산없음"
    elif 1 < budget_amount < 100000000:
        worksheet.cell(row=row, column=6).value = 0
        worksheet.cell(row=row, column=7).value = budget_amount
    elif budget_amount >= 100000000:
        worksheet.cell(row=row, column=6).value = ""
        worksheet.cell(row=row, column=7).value = budget_amount
    else:
        worksheet.cell(row=row, column=6).value = "?"
        worksheet.cell(row=row, column=9).value = budget_amount
    
    
    
    organization_list = ['한국토지주택공사', '방위사업청', '한국마사회', '한국수자원공사', '한국가스공사', '한국도로공사', '해군군수사령부',
                         '육군군수사령부', '공군군수사령부', '(주)유니비즈컨설팅']
    if organization_name in organization_list:
        worksheet.cell(row=row, column=6).value = 0

"""

def process_row(row, worksheet):
    bf_spec_no = worksheet.cell(row=row, column=11).value
    data = fetch_bf_spec_data(bf_spec_no)

    if data:
        dlBfSpecM = data.get("dlBfSpecM", {})
        budget_amount = dlBfSpecM.get("alotBgtPrspAmt", "N/A")  # 배정예산액 (부가세 포함)
    else:
        print(f"Row {row}: No data extracted for bfSpecNo {bf_spec_no}")

    if budget_amount is None:
        worksheet.cell(row=row, column=6).value = "?"
        worksheet.cell(row=row, column=7).value = "예산없음"
    elif 1 < budget_amount < 100000000:
        worksheet.cell(row=row, column=6).value = 0
        worksheet.cell(row=row, column=7).value = budget_amount
    elif budget_amount >= 100000000:
        worksheet.cell(row=row, column=6).value = ""
        worksheet.cell(row=row, column=7).value = budget_amount
    else:
        worksheet.cell(row=row, column=6).value = "?"
        worksheet.cell(row=row, column=9).value = budget_amount

    organization_list = ['한국토지주택공사', '방위사업청', '한국마사회', '한국수자원공사', '한국가스공사', '한국도로공사', '해군군수사령부',
                         '육군군수사령부', '공군군수사령부', '(주)유니비즈컨설팅']
    if organization_name in organization_list:
        worksheet.cell(row=row, column=6).value = 0
# %%
excel_file_path = input("파일명+확장자: ")
workbook = openpyxl.load_workbook(excel_file_path)
worksheet = workbook['상세정보_작업']

with concurrent.futures.ThreadPoolExecutor() as executor:
    row = 76
    while True:
        organization_name = worksheet.cell(row=row, column=14).value
        if organization_name is None:
            break
        executor.submit(process_row, row, worksheet)
        row += 1

# %%
workbook.save(excel_file_path)
print("\n완료")