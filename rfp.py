# %%
import os
from datetime import datetime, timedelta
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from getpass import getpass
import pandas as pd
from glob import glob
import xlrd
import time
from concurrent.futures import ThreadPoolExecutor
import pyautogui
import openpyxl
from openpyxl import load_workbook
import subprocess

# %%
def get_dates():
    today = datetime.now()
    weekday = today.weekday()
    date_list = []

    try:
        if weekday == 1:
            date_list = [
            f"{(today - timedelta(days=i)).year}.{(today - timedelta(days=i)).month}.{(today - timedelta(days=i)).day}"
            for i in range(5, 0, -1)
        ]
        
        elif weekday == 3:
            date_list = [
            f"{(today - timedelta(days=i)).year}.{(today - timedelta(days=i)).month}.{(today - timedelta(days=i)).day}"
            for i in range(2, 0, -1)
        ]

        else:
            raise ValueError("화요일이나 목요일이 아님")

    except Exception as e:
        print(f"오류 발생: {e}")

    return date_list

# %%
def get_urls(date_list):
    urls = []

    for date in date_list:
        url_first = f"https://www.kbid.co.kr/common/main_search_result.htm?lstFindList=1&Desc=desc&GetUp=336110%2C336100%7C%EC%82%AC%EB%AC%B4%EC%9A%A9%EA%B8%B0%EA%B8%B0%EB%B0%8F%EB%B3%B4%EC%A1%B0%EC%9A%A9%ED%92%88%2F%EC%BB%B4%ED%93%A8%ED%84%B0+%EC%86%8C%EB%AA%A8%ED%92%88+%EC%84%A4%EB%B9%84&GetTname=I&GetArea=&GetSArea=&Kind_type=0&FindG2b=&FindG2bFull=&Tname=I&lstWork=336110%2C336100&lstKind=&lstArea=999999&lstSArea=&rdoFindDate=1&txtSDate={date}&txtEDate={date}&rdoFindWord=1&txtFindWord=&gCode=&lstFindList=1&lstViewList=100"
        url_second = f"https://www.kbid.co.kr/common/main_search_result.htm?lstFindList=1&Desc=desc&GetUp=210110%2C222999%2C222110%2C222130%2C222120%7C%EC%A0%84%EC%9E%90.%EC%A0%95%EB%B3%B4+ENG.%2F%EA%B8%B0%ED%83%80%EC%A0%95%EB%B3%B4%ED%86%B5%EC%8B%A0%EA%B4%80%EB%A0%A8%2F%EC%86%8C%ED%94%84%ED%8A%B8%EC%9B%A8%EC%96%B4%EC%82%AC%EC%97%85%EC%9E%90%2FSI%2F%EC%A0%95%EB%B3%B4%EB%B3%B4%ED%98%B8%2F%EC%BB%B4%ED%93%A8%ED%84%B0%EB%B0%8F%EC%A3%BC%EB%B3%80%EA%B8%B0%EA%B8%B0%EC%9C%A0%EC%A7%80%EB%B3%B4%EC%88%98&GetTname=Y&GetArea=&GetSArea=&Kind_type=0&FindG2b=&FindG2bFull=&Tname=Y&lstWork=210110%2C222999%2C222110%2C222130%2C222120&lstKind=&lstArea=999999&lstSArea=&rdoFindDate=1&txtSDate={date}&txtEDate={date}&rdoFindWord=1&txtFindWord=&gCode=&lstFindList=1&lstViewList=100"
        
        urls.append((url_first, date))
        urls.append((url_second, date))

    return urls

# %%
def login(driver, username, password):
    driver.get("https://www.kbid.co.kr/login/common_login.htm")
    try:
        id = driver.find_element(By.XPATH, '//*[@id="MemID"]')
        id.click()
        id.send_keys(username)

        pw = driver.find_element(By.XPATH, '//*[@id="MemPW"]')
        pw.click()
        pw.send_keys(password)

        button = driver.find_element(By.XPATH, '//*[@id="FLogin"]/fieldset/ul/li[3]/input')
        button.click()

        WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CSS_SELECTOR, "selector_for_logged_in_element")))  # 로그인 성공을 확인할 요소의 CSS 선택자 입력
    except Exception as e:
        print(f"로그인 중 오류 발생: {e}")
        

# %%
def update_first_column(download_path, file_date):
    files = glob(f'{download_path}/*.xls')
    
    latest_file =  max(files, key=os.path.getctime)
    file_name = os.path.splitext(os.path.basename(latest_file))[0] + '.xlsx'
    latest_file = xlrd.open_workbook(latest_file, ignore_workbook_corruption=True)
    
    temp = pd.read_excel(latest_file)
    temp.iloc[:, 0] = file_date
    temp.to_excel(f'{download_path}/{file_name}', index=False)

# %%
def process_urls(driver, urls, download_path):
    driver.implicitly_wait(3)
    
    for url, file_date in urls:
        driver.get(url)
        
        for page_number in range(1, 4):
            try:
                try:
                    page_button = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, f"//div[@class='list-page']//a[contains(text(), '{page_number}')]")))
                    page_button.click()
                except Exception:
                    continue

                all_check = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.list-board > div.list-head > span > input#allCheck")))
                all_check.click()

                excel_button = driver.find_element(By.CSS_SELECTOR, "div.list-function > input.btn-list-function.print-excel")
                excel_button.click()
                
                time.sleep(5)
                
                updates = update_first_column(download_path, file_date)
                updates
            
            except Exception as e:
                continue

# %%
def merge_excel_files(path, output_file):
    total = pd.DataFrame()
    
    excel_files = glob(f'{path}/*.xlsx')
    excel_files.sort()
    
    for file in excel_files:
        try:
            temp = pd.read_excel(file)
            total = pd.concat([total, temp], ignore_index=False)
        except Exception as e:
            print(f"파일 {file} 처리 중 오류 발생: {e}")

    del total['수요기관']
    del total['입찰개시일']
    
    num_columns = total.shape[0]
    total.to_excel(output_file, index=False, header=False)
    
    return num_columns

# %%
def update_sheet(file_name, date):
    df = pd.read_excel(file_name)
    origin_file = 'KBID_RFP_수집_2024mmdd(mmdd).xlsx'
    result_file = f'KBID_RFP_수집_2024{date}(mmdd).xlsx'
    
    with pd.ExcelWriter(origin_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='tmp_1', index=False)
    
    workbook = load_workbook(origin_file)
    workbook.save(result_file)
    
    return result_file

# %%
def update_row(row, ws):
    period = ws.cell(row=row, column=21).value
    price = ws.cell(row=row, column=17).value
    result = ws.cell(row=row, column=22)

    if period > -10:
        result.value = 0
    if 100 < price < 90909091:
        result.value = 0
        

def update_excel(num_columns, file_name):
    if os.name == 'nt':  # Windows
        os.startfile(file_path)
    elif os.name == 'posix':  # macOS
        subprocess.run(['open', file_name])

    time.sleep(5)  # 엑셀 파일이 열릴 때까지 대기

    # 파일 저장 (macOS에서는 'command' 키 사용)
    if os.name == 'nt':  # Windows
        with pyautogui.hold('ctrl'):
            time.sleep(1)
            pyautogui.press('s')
    else:  # macOS
        with pyautogui.hold('fn'):
            time.sleep(1)
            pyautogui.press('f9')
        
        time.sleep(1)
        
        with pyautogui.hold('command'):
            time.sleep(1)
            pyautogui.press('s')

    time.sleep(1)

    # 엑셀 종료 (macOS에서는 'command' 키 사용)
    if os.name == 'nt':  # Windows
        with pyautogui.hold('ctrl'):
            time.sleep(1)
            pyautogui.press('w')
    else:  # macOS
        with pyautogui.hold('command'):
            time.sleep(1)
            pyautogui.press('w')

    time.sleep(2)  # 엑셀 종료 대기
    
    workbook = openpyxl.load_workbook(file_name, data_only=True)
    ws = workbook['상세정보_작업']
    time.sleep(1)
    
    # ThreadPoolExecutor를 사용하여 병렬로 행 업데이트
    with ThreadPoolExecutor() as executor:
        executor.map(lambda row: update_row(row, ws), range(2, num_columns + 1))

    workbook.save(file_name)

# %%
def main():
    today = datetime.strftime(datetime.now(), "%m%d")
    dates = get_dates()
    urls = get_urls(dates)
    username = "panda0070"
    password = getpass("password: ")

    chromedriver_autoinstaller.install()
    
    download_path = os.getcwd()+f'/{today}'
    os.makedirs(download_path) 
    
    options = webdriver.ChromeOptions()
    options.add_experimental_option('prefs', {
        "download.default_directory": download_path,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    })

    driver = webdriver.Chrome(options=options)

    login(driver, username, password)
    process_urls(driver, urls, download_path)

    driver.quit()
    
    file_name = f'{today}.xlsx'
    
    num_columns = merge_excel_files(download_path, file_name)
    result_file = update_sheet(file_name, today)
    update_excel(num_columns, result_file)
    
    print("\n완료")

if __name__ == "__main__":
    main()


