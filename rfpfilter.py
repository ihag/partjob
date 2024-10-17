# %%
import os
import openpyxl
import pyautogui
from bs4 import BeautifulSoup
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# %%
DOWNLOAD_SLEEP_TIME = 5
SAVE_WAIT_TIME = 1

# %%
# Step 1
def open_excel_file(file_path):
    if os.path.exists(file_path):
        return file_path
    else:
        raise FileNotFoundError(f"File not found: {file_path}")

# %%
# Step 2: V열의 값이 없고, T열의 값이 '??'가 아닌 경우 처리
def process_excel(file_path, driver):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['상세정보_작업']
    
    trigger_excel_calculation()

    # Step 2: V열의 값이 없고, T열의 값이 '??'가 아닌 경우 처리
    for row in range(2, sheet.max_row + 1):  # 데이터가 2행부터 있을 때
        t_value = sheet.cell(row=row, column=20).value  # T열
        
        # T열의 값이 없으면 바로 다음 행으로 넘어감
        if not t_value or t_value == '??':
            continue  # T열 값이 없거나 '??'일 경우 해당 행 스킵

        v_value = sheet.cell(row=row, column=22).value  # V열

        # 기본 조건: V열에 값이 없을 때만 진행
        if v_value is None:
            hyperlink = sheet.cell(row=row, column=20).hyperlink
            if hyperlink:
                url = hyperlink.target
                number = get_registration_number(url, driver)

                # Step 3: J열에서 같은 값을 찾기
                if number and number != '0':
                    match_in_J = check_number_in_J_column_set(sheet, number)
                    if match_in_J:
                        sheet.cell(row=row, column=22).value = 0  # V열 (22번째 열)에 0 입력

    wb.save(file_path)

# %%
# Step 3: 엑셀 계산 실행
def trigger_excel_calculation():
    if os.name == 'nt':  # Windows
        with pyautogui.hold('ctrl'):
            time.sleep(SAVE_WAIT_TIME)
            pyautogui.press('s')
    else:  # macOS
        with pyautogui.hold('fn'):
            time.sleep(SAVE_WAIT_TIME)
            pyautogui.press('f9')
        
        time.sleep(1)
        
        with pyautogui.hold('command'):
            time.sleep(SAVE_WAIT_TIME)
            pyautogui.press('s')

    time.sleep(SAVE_WAIT_TIME)

    if os.name == 'nt':  # Windows
        with pyautogui.hold('ctrl'):
            time.sleep(SAVE_WAIT_TIME)
            pyautogui.press('w')
    else:  # macOS
        with pyautogui.hold('command'):
            time.sleep(SAVE_WAIT_TIME)
            pyautogui.press('w')

    time.sleep(SAVE_WAIT_TIME)

# %%
# Step 4: URL에서 사전규격등록번호 가져오기
def get_registration_number(url, driver):
    # driver.get(url)
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    reg_no = int(soup.find(string="사전규격등록번호").find_next("td").text.strip())
    
    return reg_no

    """
    try:
        number_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//td[@class="info"]/span[@id="specRegNo"]'))  # 정확한 XPath를 사용해야 함
        )
        number = number_element.text.strip()  # 숫자 가져오기
        return number
    except Exception as e:
        print(f"Error fetching number from URL {url}: {e}")
        return None
    """

# %%
# Step 5: J열에서 같은 값을 찾기
def check_number_in_J_column_set(sheet, number):
    j_values_set = {sheet.cell(row=row, column=10).value for row in range(2, sheet.max_row + 1)}  # J열은 10번째 열
    
    return number in j_values_set

# %%
# Step 6: 웹드라이버 설정 및 실행
def setup_webdriver():
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)
    return driver

# %%
# Step 7: 현재 폴더에서 파일명을 입력받아 엑셀 파일을 처리
def main():
    current_folder = os.getcwd()
    
    file_name = input("Enter the Excel file name (with extension): ")
    file_path = os.path.join(current_folder, file_name)
    
    try:
        excel_file = open_excel_file(file_path)
    except FileNotFoundError as e:
        print(e)
        return

    driver = setup_webdriver()

    try:
        process_excel(excel_file, driver)
    finally:
        driver.quit()

    print("Completed.")

if __name__ == "__main__":
    main()
