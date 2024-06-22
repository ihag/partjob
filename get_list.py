# %%
import concurrent.futures
import openpyxl
import requests
import re
import time
import sys
from bs4 import BeautifulSoup

# %%
def get_list(url_value):
    url = f"https://www.g2b.go.kr:8081/ep/preparation/prestd/preStdDtl.do?preStdRegNo={url_value}"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # 1. 공개일시
    open_date = soup.find(string='공개일시').find_next("td").text.strip()
    # 2. 의견등록마감일시
    opinion_deadline = soup.find(string="의견등록마감일시").find_next("td").text.strip()
    # 3. 사전규격등록번호
    reg_no = int(soup.find(string="사전규격등록번호").find_next("td").text.strip())
    # 4. 참조번호
    ref_no = soup.find(string="참조번호").find_next("td").text.strip()
    # 5. 사업명
    project_name = soup.find("td", {"colspan": "3"}).find("div", class_="tb_inner").text.strip()
    # 6. 수요기관
    demand_agency = soup.find(string="수요기관").find_next("td").text.strip()
    # 7. 배정예산액
    budget_text = soup.find("th", string="배정예산액").find_next("td").text.strip()
    budget_amount = int(re.sub(r'[^0-9]', '', budget_text))

    return {
                "open_date": open_date,
                "opinion_deadline": opinion_deadline,
                "reg_no": reg_no,
                "ref_no": ref_no,
                "project_name": project_name,
                "demand_agency": demand_agency,
                "budget_amount": budget_amount
                }

# %%
def process_row(row,list_row, worksheet):
    url_value = worksheet.cell(row=row, column=11).value
    
    data_dict = get_list(url_value)
    
    worksheet.cell(row=list_row, column=1).value = list_row-1
    worksheet.cell(row=list_row, column=5).value = data_dict["open_date"][:10]
    worksheet.cell(row=list_row, column=6).value = data_dict["opinion_deadline"][:10]
    worksheet.cell(row=list_row, column=10).value = data_dict["reg_no"]
    worksheet.cell(row=list_row, column=11).value = data_dict["ref_no"]
    worksheet.cell(row=list_row, column=12).value = data_dict["project_name"]
    worksheet.cell(row=list_row, column=13).value = data_dict["demand_agency"]
    worksheet.cell(row=list_row, column=15).value = data_dict["project_name"]
    worksheet.cell(row=list_row, column=17).value = data_dict["budget_amount"]

# %%
excel_file_path = input("파일명+확장자: ")
workbook = openpyxl.load_workbook(excel_file_path)
worksheet = workbook['상세정보_작업']

with concurrent.futures.ThreadPoolExecutor() as executor:
    row = 76
    list_row = 2
    while True:
        cell_value = worksheet.cell(row=row, column=6).value
        if cell_value != 1:
            break
        executor.submit(process_row, row, list_row, worksheet)
        row += 1
        list_row += 1

# %%
workbook.save(excel_file_path)

# %%
def progress_bar(current, total, bar_length=50):
    percent = float(current) / total
    hashes = '#' * int(percent * bar_length)
    spaces = '-' * (bar_length - len(hashes))
    sys.stdout.write("\r[{0}] {1}%".format(hashes + spaces, int(percent * 100)))
    sys.stdout.flush()

total_steps = 100

for i in range(total_steps + 1):
    progress_bar(i, total_steps)
    time.sleep(0.1)

print("\n완료")


