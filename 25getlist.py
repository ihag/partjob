import requests
import html
from openpyxl import load_workbook
from datetime import datetime

def fetch_bf_spec_data(bf_spec_no):
    url = "https://www.g2b.go.kr/pn/pnz/pnza/BfSpec/selectBfSpec.do"
    
    payload = {
        "dlParamM": {
            "bfSpecRegNo": bf_spec_no,
            "prcmBsneSeCd": "",  # 필요에 따라 수정
            "opnnSqno": "",      # 필요에 따라 수정
            "jobType": ""        # 필요에 따라 수정
        }
    }

    headers = {
        'Content-Type': 'application/json;charset=UTF-8',
        'Accept': 'application/json',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        'Referer': 'https://www.g2b.go.kr/',
        'Origin': 'https://www.g2b.go.kr',
        'Cookie': 'WHATAP=z1c19kr5h17kv5; XTVID=A2501060840080519; JSESSIONID=YWE4OTU3ZjItMDhmMy00MjZjLTgzMmMtMzEzZjM1ZWM0N2Qw;',  # 최신 쿠키 입력
        'Menu-Info': '{"menuNo":"01141","menuCangVal":"PRVA004_02","bsneClsfCd":"%EC%97%85130025","scrnNo":"05929"}',
        'SubmissionId': 'mf_wfm_container_sbmSrch',
        'Priority': 'u=1, i',
        'Sec-CH-UA': '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
        'Sec-CH-UA-Mobile': '?0',
        'Sec-CH-UA-Platform': '"Windows"',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
    }

    response = requests.post(url, headers=headers, json=payload)

    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error: {response.status_code} {response.text}")
        return None

def process_row(row, list_row, worksheet):
    bf_spec_no = worksheet.cell(row=row, column=11).value
    data = fetch_bf_spec_data(bf_spec_no)

    if data:
        # 데이터 기록
        worksheet.cell(row=list_row, column=1).value = list_row - 1
        worksheet.cell(row=list_row, column=3).value = datetime.now().date()
        
        # 필요한 정보 추출
        dlBfSpecM = data.get("dlBfSpecM", {})
        worksheet.cell(row=list_row, column=5).value = dlBfSpecM.get("specRlsDt", "N/A").strip().replace('/', '-')[:10]  # 사전규격 공개일시
        worksheet.cell(row=list_row, column=6).value = dlBfSpecM.get("opnnRegDdlnDt", "N/A").strip().replace('/', '-')[:10]  # 의견등록마감일시
        worksheet.cell(row=list_row, column=10).value = dlBfSpecM.get("bfSpecRegNo", "N/A")  # 사전규격등록번호
        worksheet.cell(row=list_row, column=11).value = dlBfSpecM.get("oderPlanNo", "N/A")  # 발주계획통합번호
        bf_spec_name = dlBfSpecM.get("bfSpecNm", "N/A") # 사전규격명
        worksheet.cell(row=list_row, column=12).value = html.unescape(bf_spec_name)  # HTML 엔티티 변환
        worksheet.cell(row=list_row, column=15).value = html.unescape(bf_spec_name)  # HTML 엔티티 변환
        dmst_unty_groupname = dlBfSpecM.get("dmstUntyGrpNm", "N/A") # 수요기관
        worksheet.cell(row=list_row, column=13).value = html.unescape(dmst_unty_groupname)  # HTML 엔티티 변환
        worksheet.cell(row=list_row, column=17).value = dlBfSpecM.get("alotBgtPrspAmt", "N/A")  # 배정예산액 (부가세 포함)
        print(f"Row {list_row}: Data written successfully -> {dlBfSpecM}")
    else:
        print(f"Row {list_row}: No data extracted for bfSpecNo {bf_spec_no}")

def main():
    excel_file_path = input("파일명+확장자: ")
    workbook = load_workbook(excel_file_path)
    worksheet = workbook['상세정보_작업']

    row = 76  # 데이터 시작 행
    list_row = 2

    while True:
        cell_value = worksheet.cell(row=row, column=6).value
        if cell_value != 1:  # 특정 값이 1인 셀만 처리
            break
        process_row(row, list_row, worksheet)
        row += 1
        list_row += 1

    # 엑셀 저장
    workbook.save(excel_file_path)
    print("\n완료")

if __name__ == "__main__":
    main()
