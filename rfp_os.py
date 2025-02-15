# %%
import os
from datetime import datetime, timedelta
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from getpass import getpass
import pandas as pd
from glob import glob
import xlrd
import time
from concurrent.futures import ThreadPoolExecutor
import pyautogui
import openpyxl
from openpyxl import load_workbook
import subprocess

# %%
DOWNLOAD_SLEEP_TIME = 5
SAVE_WAIT_TIME = 1

# %%
# Get the dates based on weekday (Tuesday/Thursday)
def get_dates():
    today = datetime.now()
    weekday = today.weekday()
    
    try:
        if weekday == 1:  # Tuesday
            date_list = [(today - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(5, 0, -1)]
        elif weekday == 3:  # Thursday
            date_list = [(today - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(2, 0, -1)]
        else:
            raise ValueError("The current day is not Tuesday or Thursday.")
    except Exception as e:
        print(f"Error occurred: {e}")
        date_list = []

    return date_list

# %%
# Generate URLs based on the provided dates
def get_urls(date_list):
    urls = []
    for date in date_list:
        url_first = f"https://www.kbid.co.kr/common/main_search_result.htm?lstFindList=1&Desc=desc&GetUp=336110%2C336100%7C%EC%82%AC%EB%AC%B4%EC%9A%A9%EA%B8%B0%EA%B8%B0%EB%B0%8F%EB%B3%B4%EC%A1%B0%EC%9A%A9%ED%92%88%2F%EC%BB%B4%ED%93%A8%ED%84%B0+%EC%86%8C%EB%AA%A8%ED%92%88+%EC%84%A4%EB%B9%84&GetTname=I&GetArea=&GetSArea=&Kind_type=0&FindG2b=&FindG2bFull=&Tname=I&lstWork=336110%2C336100&lstKind=&lstArea=999999&lstSArea=&rdoFindDate=1&txtSDate={date}&txtEDate={date}&rdoFindWord=1&txtFindWord=&gCode=&lstFindList=1&lstViewList=100"
        url_second = f"https://www.kbid.co.kr/common/main_search_result.htm?lstFindList=1&Desc=desc&GetUp=210110%2C222999%2C222110%2C222130%2C222120%7C%EC%A0%84%EC%9E%90.%EC%A0%95%EB%B3%B4+ENG.%2F%EA%B8%B0%ED%83%80%EC%A0%95%EB%B3%B4%ED%86%B5%EC%8B%A0%EA%B4%80%EB%A0%A8%2F%EC%86%8C%ED%94%84%ED%8A%B8%EC%9B%A8%EC%96%B4%EC%82%AC%EC%97%85%EC%9E%90%2FSI%2F%EC%A0%95%EB%B3%B4%EB%B3%B4%ED%98%B8%2F%EC%BB%B4%ED%93%A8%ED%84%B0%EB%B0%8F%EC%A3%BC%EB%B3%80%EA%B8%B0%EA%B8%B0%EC%9C%A0%EC%A7%80%EB%B3%B4%EC%88%98&GetTname=Y&GetArea=&GetSArea=&Kind_type=0&FindG2b=&FindG2bFull=&Tname=Y&lstWork=210110%2C222999%2C222110%2C222130%2C222120&lstKind=&lstArea=999999&lstSArea=&rdoFindDate=1&txtSDate={date}&txtEDate={date}&rdoFindWord=1&txtFindWord=&gCode=&lstFindList=1&lstViewList=100"
        
        urls.append((url_first, date))
        urls.append((url_second, date))
    return urls

# %%
# Log into the website
def login(driver, username, password):
    driver.get("https://www.kbid.co.kr/login/common_login.htm")
    try:
        id_box = driver.find_element(By.XPATH, '//*[@id="MemID"]')
        id_box.send_keys(username)
        
        pw_box = driver.find_element(By.XPATH, '//*[@id="MemPW"]')
        pw_box.send_keys(password)
        
        login_button = driver.find_element(By.XPATH, '//*[@id="FLogin"]/fieldset/ul/li[3]/input')
        login_button.click()
        
        # Check if login is successful by waiting for a known element on the logged-in page
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "selector_for_logged_in_element")))
    except Exception as e:
        print(f"Login failed: {e}")

# %%
# Update the first column of the downloaded Excel file with the corresponding date
def update_first_column(download_path, file_date):
    files = glob(os.path.join(download_path, '*.xls'))
    
    if not files:
        print(f"No files found in {download_path}.")
        return
    
    latest_file = max(files, key=os.path.getctime)
    file_name = os.path.splitext(os.path.basename(latest_file))[0] + '.xlsx'
    latest_file = xlrd.open_workbook(latest_file, ignore_workbook_corruption=True)

    # Open and update the Excel file
   
    temp_df = pd.read_excel(latest_file)
    temp_df.iloc[:, 0] = file_date  # Update first column with the date
    temp_df.to_excel(os.path.join(download_path, file_name), index=False)


# %%
# Process the URLs and download the relevant files
def process_urls(driver, urls, download_path):
    driver.implicitly_wait(3)
    for url, file_date in urls:
        driver.get(url)
        

        for page_number in range(1, 4):
            try:
                try:
                    page_button = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, f"//div[@class='list-page']//a[contains(text(), '{page_number}')]")))
                    page_button.click()
                except Exception:
                    continue

                all_check = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.list-board > div.list-head > span > input#allCheck")))
                all_check.click()

                excel_button = driver.find_element(By.CSS_SELECTOR, "div.list-function > input.btn-list-function.print-excel")
                excel_button.click()
                
                time.sleep(DOWNLOAD_SLEEP_TIME)
                
                updates = update_first_column(download_path, file_date)
                updates
            
            except Exception as e:
                continue

# %%
# Merge downloaded Excel files into a single file
def merge_excel_files(path, output_file):
    total_df = pd.DataFrame()
    excel_files = glob(os.path.join(path, '*.xlsx'))
    
    for file in excel_files:
        try:
            temp_df = pd.read_excel(file)
            total_df = pd.concat([total_df, temp_df], ignore_index=True)
        except Exception as e:
            print(f"Error reading file {file}: {e}")
    
    # Remove unwanted columns
    del total_df['수요기관']
    del total_df['입찰개시일']
    
    total_df.to_excel(output_file, index=False, header=False)
    return total_df.shape[0]

# %%
# Update a specific Excel file by modifying cells based on conditions
def update_row(row, ws):
    period = ws.cell(row=row, column=21).value
    price = ws.cell(row=row, column=17).value
    result = ws.cell(row=row, column=22)
    
    if period > -10:
        result.value = 0
    if 100 < price < 90909091:
        result.value = 0

# %%
def update_excel(num_columns, file_name):
    # Open Excel file based on OS
    if os.name == 'nt':  # Windows
        os.startfile(file_name)
    elif os.name == 'posix':  # macOS
        subprocess.run(['open', file_name])

    time.sleep(SAVE_WAIT_TIME)
    
    trigger_excel_calculation()
    
    workbook = openpyxl.load_workbook(file_name, data_only=True)
    ws = workbook['상세정보_작업_2']
    
    with ThreadPoolExecutor() as executor:
        executor.map(lambda row: update_row(row, ws), range(2, num_columns + 1))
    
    workbook.save(file_name)
    
# %%
def trigger_excel_calculation():
    if os.name == 'nt':  # Windows
        with pyautogui.hold('ctrl'):
            time.sleep(SAVE_WAIT_TIME)
            pyautogui.press('s')
    else:  # macOS
        with pyautogui.hold('fn'):
            time.sleep(SAVE_WAIT_TIME)
            pyautogui.press('f9')
        
        time.sleep(1)
        
        with pyautogui.hold('command'):
            time.sleep(SAVE_WAIT_TIME)
            pyautogui.press('s')

    time.sleep(1)

    if os.name == 'nt':  # Windows
        with pyautogui.hold('ctrl'):
            time.sleep(SAVE_WAIT_TIME)
            pyautogui.press('w')
    else:  # macOS
        with pyautogui.hold('command'):
            time.sleep(SAVE_WAIT_TIME)
            pyautogui.press('w')

    time.sleep(SAVE_WAIT_TIME)

# %%
# Update the sheet with the new data from merged files
def update_sheet(merged_file_name, date):
    # Load the merged file into a DataFrame
    df = pd.read_excel(merged_file_name)
    
    # Define the origin file template and the result file name with the current date
    origin_file = 'KBID_RFP_수집_2024mmdd(mmdd).xlsx'
    result_file = f'KBID_RFP_수집_2024{date}(mmdd).xlsx'
    
    # Append the new data (from df) to the "tmp_1" sheet in the origin_file
    try:
        with pd.ExcelWriter(origin_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='tmp_1', index=False)
        
        # Save the modified file as the result file
        workbook = load_workbook(origin_file)
        workbook.save(result_file)
        
    except Exception as e:
        print(f"Error updating sheet: {e}")
    
    return result_file

# %%
def main():
    today = datetime.strftime(datetime.now(), "%m%d")
    download_path = os.path.join(os.getcwd(), today)
    os.makedirs(download_path, exist_ok=True)  # Ensure download directory exists

    dates = get_dates()
    if not dates:
        print("No valid dates found for processing.")
        return
    
    urls = get_urls(dates)
    username = input("id: ")
    password = getpass("password: ")

    chromedriver_autoinstaller.install()

    options = webdriver.ChromeOptions()
    options.add_experimental_option('prefs', {
        "download.default_directory": download_path,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    })

    driver = webdriver.Chrome(options=options)

    login(driver, username, password)
    process_urls(driver, urls, download_path)

    driver.quit()

    file_name = f'{today}.xlsx'
    
    num_columns = merge_excel_files(download_path, file_name)
    
    """
    if num_columns > 0:
        result_file = update_sheet(file_name, today)  # Use the updated function
        update_excel(num_columns, result_file)
    """
    
    print("\nCompleted")
    

if __name__ == "__main__":
    main()

# %%
