# %%
import concurrent.futures
import openpyxl
import requests
import re
from bs4 import BeautifulSoup

# %%
def extract_budget_amount(url_value):
    url = f"https://www.g2b.go.kr:8081/ep/preparation/prestd/preStdDtl.do?preStdRegNo={url_value}"
    try:
        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'html.parser')
        budget_text = soup.find(string='배정예산액').find_next('td').text.strip()
        if budget_text:
            budget_amount = int(re.sub(r'[^0-9]', '', budget_text))
            return budget_amount
        else:
            return None
    except requests.exceptions.RequestException:
        return None
    except (AttributeError, ValueError):
        return None

# %%
def process_row(row, worksheet):
    url_value = worksheet.cell(row=row, column=11).value
    organization_name = worksheet.cell(row=row, column=14).value
    
    if organization_name is None:
        return

    budget_amount = extract_budget_amount(url_value)
    
    if budget_amount is None:
        worksheet.cell(row=row, column=6).value = "?"
        worksheet.cell(row=row, column=7).value = "예산없음"
    elif 1 < budget_amount < 90000000:
        worksheet.cell(row=row, column=6).value = 0
        worksheet.cell(row=row, column=7).value = budget_amount
    elif budget_amount >= 90000000:
        worksheet.cell(row=row, column=6).value = ""
        worksheet.cell(row=row, column=7).value = budget_amount
    else:
        worksheet.cell(row=row, column=6).value = "?"
        worksheet.cell(row=row, column=9).value = budget_amount
    
    organization_list = ['한국토지주택공사', '방위사업청', '한국마사회', '한국수자원공사', '한국가스공사', '한국도로공사', '해군군수사령부',
                         '육군군수사령부', '공군군수사령부', '(주)유니비즈컨설팅']
    if organization_name in organization_list:
        worksheet.cell(row=row, column=6).value = 0

# %%
excel_file_path = input("파일명+확장자: ")
workbook = openpyxl.load_workbook(excel_file_path)
worksheet = workbook['상세정보_작업']

with concurrent.futures.ThreadPoolExecutor() as executor:
    row = 76
    while True:
        organization_name = worksheet.cell(row=row, column=14).value
        if organization_name is None:
            break
        executor.submit(process_row, row, worksheet)
        row += 1
        
    print("완료")

# %%
workbook.save(excel_file_path)

# %%
def progress_bar(current, total, bar_length=50):
    percent = float(current) / total
    hashes = '#' * int(percent * bar_length)
    spaces = '-' * (bar_length - len(hashes))
    sys.stdout.write("\r[{0}] {1}%".format(hashes + spaces, int(percent * 100)))
    sys.stdout.flush()

total_steps = 100

for i in range(total_steps + 1):
    progress_bar(i, total_steps)
    time.sleep(0.1)

print("\n완료")


